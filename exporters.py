"""
Tech-Debt-MCP 输出格式模块
支持多种输出格式：JSON, Markdown, HTML, CSV, Excel
"""
import json
import csv
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path


class OutputFormatter:
    """输出格式化器"""

    @staticmethod
    def to_json(data: Dict, pretty: bool = True) -> str:
        """转换为 JSON 格式"""
        if pretty:
            return json.dumps(data, indent=2, ensure_ascii=False)
        return json.dumps(data, ensure_ascii=False)

    @staticmethod
    def to_markdown(data: Dict, title: str = "技术债务分析报告") -> str:
        """转换为 Markdown 格式"""
        lines = [f"# {title}", ""]
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        if 'items' in data:
            lines.append("## 债务项列表")
            lines.append("")
            lines.append("| 序号 | 类型 | 文件 | 实体 | 复杂度 |")
            lines.append("|------|------|------|------|--------|")

            for i, item in enumerate(data['items'], 1):
                file_short = Path(item.get('file_path', '')).name
                lines.append(
                    f"| {i} | {item.get('type', 'N/A')} | {file_short} | "
                    f"{item.get('entity_name', 'N/A')} | {item.get('complexity', 'N/A')} |"
                )

        return "\n".join(lines)

    @staticmethod
    def to_html(data: Dict, title: str = "技术债务分析报告") -> str:
        """转换为 HTML 格式"""
        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        .high {{ color: #f44336; font-weight: bold; }}
        .medium {{ color: #ff9800; }}
        .low {{ color: #4CAF50; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
"""

        if 'items' in data:
            html += "<h2>债务项列表</h2>\n"
            html += "<table>\n"
            html += "<tr><th>序号</th><th>类型</th><th>文件</th><th>实体</th><th>复杂度</th><th>优先级</th></tr>\n"

            for i, item in enumerate(data['items'], 1):
                file_short = Path(item.get('file_path', '')).name
                complexity = item.get('complexity', 0)
                debt_score = item.get('debt_score', 0)

                priority_class = 'high' if debt_score and debt_score > 0.5 else 'medium' if debt_score and debt_score > 0.3 else 'low'

                priority_text = '高' if priority_class == 'high' else '中' if priority_class == 'medium' else '低'

                html += f"""<tr>
    <td>{i}</td>
    <td>{item.get('type', 'N/A')}</td>
    <td>{file_short}</td>
    <td>{item.get('entity_name', 'N/A')}</td>
    <td>{complexity}</td>
    <td class="{priority_class}">{priority_text}</td>
</tr>\n"""

            html += "</table>\n"

        html += "</body>\n</html>"

        return html

    @staticmethod
    def to_csv(data: Dict, filepath: str = None) -> str:
        """转换为 CSV 格式"""
        if not filepath:
            import tempfile
            filepath = tempfile.mktempfile(mode='w', suffix='.csv').name

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            if 'items' in data:
                writer = csv.DictWriter(f)
                writer.writeheader([
                    '序号', '类型', '文件', '实体', '复杂度', '债务指数', '修改次数'
                ])

                for i, item in enumerate(data['items'], 1):
                    writer.writerow({
                        '序号': i,
                        '类型': item.get('type', 'N/A'),
                        '文件': item.get('file_path', 'N/A'),
                        '实体': item.get('entity_name', 'N/A'),
                        '复杂度': item.get('complexity', 'N/A'),
                        '债务指数': item.get('debt_score', 0),
                        '修改次数': item.get('modification_frequency', 0)
                    })

        return filepath

    @staticmethod
    def to_excel(data: Dict, filepath: str = None) -> str:
        """转换为 Excel 格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Pattern

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "技术债务分析报告"

            # 添加标题
            ws['A1'] = datetime.now().strftime('%Y-%m-%d')
            ws['A1'].font = Font(size=14, bold=True)

            # 添加数据
            if 'items' in data:
                headers = ['序号', '类型', '文件', '实体', '复杂度', '债务指数', '修改次数']
                ws.append(headers)

                for i, item in enumerate(data['items'], 1):
                    row = [
                        i,
                        item.get('type', 'N/A'),
                        Path(item.get('file_path', '')).name,
                        item.get('entity_name', 'N/A'),
                        item.get('complexity', 'N/A'),
                        item.get('debt_score', 0),
                        item.get('modification_frequency', 0)
                    ]
                    ws.append(row)

            # 应用样式
            for cell in ws['A1:1']:
                cell.font = Font(bold=True)

            if not filepath:
                import tempfile
                filepath = tempfile.mktempfile(suffix='.xlsx').name

            wb.save(filepath)
            return filepath

        except ImportError:
            # 如果 openpyxl 未安装，回退到 CSV
            return OutputFormatter.to_csv(data, filepath.replace('.xlsx', '.csv'))


# 使用示例
if __name__ == "__main__":
    # 示例数据
    sample_data = {
        "items": [
            {
                "file_path": "/path/to/file.java",
                "entity_name": "complexMethod",
                "type": "COMPLEX_METHOD",
                "complexity": 25,
                "debt_score": 0.75,
                "modification_frequency": 42
            }
        ]
    }

    # 测试各种格式
    print("JSON 格式:")
    print(OutputFormatter.to_json(sample_data))

    print("\nMarkdown 格式:")
    print(OutputFormatter.to_markdown(sample_data))

    print("\nHTML 格式已保存")
    html_path = OutputFormatter.to_html(sample_data)
    print(f"文件: {html_path}")

    print("\nCSV 格式已保存")
    csv_path = OutputFormatter.to_csv(sample_data)
    print(f"文件: {csv_path}")